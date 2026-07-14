"""
Jeofizik Ders Yuku Dagilimi - Sayistay perspektifiyle "kamu zarari" riski analizi.

!!! VERI DOGRULAMA NOTU (14.07.2026 itibariyle guncellendi) !!!
Bu Excel tablosu, 13.07.2026 tarihinde Doc. Dr. Savas Karabulut tarafindan bolum e-posta listesi ve X (Twitter) uzerinden kamuya acik olarak paylasilmistir. Bolum
Baskanligi 07.07.2026 tarihli yazisinda verinin kaynagini teyit etmedigini
bildirmisti (ayrica kamu zarari/Sayistay konularini "bilgi edinme" kapsami
disinda gorerek esasa girmedi); ancak tablo, bolumun kendi resmi web
sitesinde (jeofizikmuhendislik.iuc.edu.tr/tr/content/egitim/ders-programlari)
yayinlanan, Bolum Baskani imzali 4 resmi ders programi belgesiyle tek tek
karsilastirilarak dogrulandi: ~30 ogretim uyesinden 29'unun toplam ders
sayisi birebir eslesti. Tek net uyusmazlik: Zeynep Cansu (Excel'de 1,
resmi programda 2 ders). Ancak KREDI sutunu hala "haftalik ders saati"nin
yaklasik bir proxy'sidir (asagidaki kisit notuna bakiniz) - bu script'teki
esik-alti/esik-ustu hesaplari bu proxy'ye dayanir.

Baglam: Istanbul Universitesi 2017 ve 2023 Sayistay Duzenlilik Denetim
Raporlari, bazi ogretim uyelerinin zorunlu ders yukunu tamamlamadigini,
buna karsilik bazilarina ek ders ucreti odenerek fazladan ders
yuklendigini tespit etmis ve bunu KAMU ZARARI olarak nitelendirmistir
(kaynak: aliosmanoncel.blogspot.com/2025/01/education.html).

2547 sayili Kanun m.36: Ogretim uyeleri HAFTADA EN AZ 10 SAAT ders
vermekle yukumludur. Bu esigin altinda kalanlar "ders yukunu
tamamlamamis", esigi asanlar ise potansiyel ek ders ucreti hak edenlerdir.

ONEMLI KISIT: Bu tablodaki "Kredi" sutunu resmi "haftalik ders saati"
ile birebir ayni degildir (milli kredi = teorik + uygulama/2 formulune
gore hesaplanir, ders saatinden farkli olabilir). Burada KREDI, saat
icin en yakin mevcut proxy olarak kullanilmistir; kesin haftalik saat
degerleri icin resmi ders programi/ek ders puantaj cetveli teyit
edilmelidir.
"""
import openpyxl

XLSX_PATH = "../data/Jeofizik_Ders_Yuku_Dagilimi.xlsx"
ANA_TABLO = "Ana Tablo"
ANABILIM_DALI_SHEET = "Anabim Dallarına göre"
ASGARI_HAFTALIK_SAAT = 10  # 2547 sayili Kanun m.36 (ogretim uyesi)

TR_MAP = str.maketrans({
    "İ": "i", "I": "i", "ı": "i",
    "Ş": "s", "ş": "s", "Ğ": "g", "ğ": "g",
    "Ü": "u", "ü": "u", "Ö": "o", "ö": "o", "Ç": "c", "ç": "c",
})


def normalize(name):
    name = name.replace("Dr. Öğr. Üyesi (Doç. Dr.)", "")
    words = name.translate(TR_MAP).split()
    return words[-1].lower() if words else ""


wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

dali_ws = wb[ANABILIM_DALI_SHEET]
resmi_uyeler = {}
for row in dali_ws.iter_rows(min_row=9, max_row=30, values_only=True):
    anabilim_dali, isim, _unvan = row[0], row[1], row[2]
    if anabilim_dali and isim:
        resmi_uyeler[normalize(isim)] = anabilim_dali

ws = wb[ANA_TABLO]
headers = [c.value for c in ws[1]]
idx_name = headers.index("Öğretim Üyesi")
idx_unvan = headers.index("Ünvan")

# Guz donemi kredi sutunlari: Lisans Z, Lisans S, YL Z, YL S
idx_guz = [headers.index("Güz Lisans Z Kredi"), headers.index("Güz Lisans S Kredi"),
           headers.index("Güz YL/Doktora Z Kredi"), headers.index("Güz YL/Doktora S Kredi")]
# Bahar donemi kredi sutunlari
idx_bahar = [headers.index("Bahar Lisans Z Kredi"), headers.index("Bahar Lisans S Kredi"),
             headers.index("Bahar YL/Doktora Z Kredi"), headers.index("Bahar YL/Doktora S Kredi")]

kayitlar = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[idx_name]
    if not name:
        continue
    key = normalize(name)
    if key not in resmi_uyeler:
        continue
    guz_kredi = sum(row[i] or 0 for i in idx_guz)
    bahar_kredi = sum(row[i] or 0 for i in idx_bahar)
    kayitlar.append({
        "isim": name, "unvan": row[idx_unvan], "dali": resmi_uyeler[key],
        "guz": guz_kredi, "bahar": bahar_kredi,
    })

print("=" * 90)
print("SAYISTAY PERSPEKTIFI: DONEM BAZLI KREDI YUKU vs 10 SAAT/HAFTA ASGARI ESIGI")
print("(Kredi, haftalik ders saatine en yakin proxy - kesin degil, bkz. dosya basi not)")
print("=" * 90)
print(f"{'Öğretim Üyesi':35s} {'Ünvan':16s} {'Güz':>6s} {'Bahar':>6s}  Durum")
print("-" * 90)

esik_alti = []
esik_ustu = []
for k in sorted(kayitlar, key=lambda x: -(x["guz"] + x["bahar"])):
    durum_bits = []
    if k["guz"] < ASGARI_HAFTALIK_SAAT:
        durum_bits.append(f"GÜZ ALTI ({k['guz']:.1f}<{ASGARI_HAFTALIK_SAAT})")
        esik_alti.append((k, "Güz"))
    elif k["guz"] > ASGARI_HAFTALIK_SAAT:
        durum_bits.append(f"GÜZ FAZLA ({k['guz']:.1f}>{ASGARI_HAFTALIK_SAAT}, ek ders ücreti riski)")
        esik_ustu.append((k, "Güz"))
    if k["bahar"] < ASGARI_HAFTALIK_SAAT:
        durum_bits.append(f"BAHAR ALTI ({k['bahar']:.1f}<{ASGARI_HAFTALIK_SAAT})")
        esik_alti.append((k, "Bahar"))
    elif k["bahar"] > ASGARI_HAFTALIK_SAAT:
        durum_bits.append(f"BAHAR FAZLA ({k['bahar']:.1f}>{ASGARI_HAFTALIK_SAAT}, ek ders ücreti riski)")
        esik_ustu.append((k, "Bahar"))
    print(f"{k['isim']:35s} {k['unvan']:16s} {k['guz']:6.1f} {k['bahar']:6.1f}  {'; '.join(durum_bits)}")

print("\n" + "=" * 90)
print(f"ESIGIN ALTINDA KALAN DONEM SAYISI (tamamlanmamis ders yuku): {len(esik_alti)}")
print(f"ESIGI ASAN DONEM SAYISI (potansiyel ek ders ücreti): {len(esik_ustu)}")
print("\nSayistay'in tespit ettigi 'kamu zarari' senaryosu tam olarak budur: ayni")
print("bolumde, ayni donemde, esigin ALTINDA kalanlar VARKEN, esigi ASANLARA")
print("ek ders ücreti odenmesi. Esik alti olan ogretim uyelerine once mevcut")
print("zorunlu/secmeli ders atanmasi (esigi asanlarin fazlasindan devredilerek)")
print("gerekirdi; bu yapilmadan ek ders ücreti odenmesi mevzuata aykiridir.")
