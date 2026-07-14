"""
Jeofizik Ders Yuku Dagilimi - Prof. Dr. unvanli ogretim uyeleri uzerinde
zorunlu ders yuku analizi.

!!! VERI DOGRULAMA NOTU (14.07.2026 itibariyle guncellendi) !!!
Bu Excel tablosu, 13.07.2026 tarihinde Doc. Dr. Savas Karabulut tarafindan bolum e-posta listesi ve X (Twitter) uzerinden kamuya acik olarak paylasilmistir. Bolum
Baskanligi 07.07.2026 tarihli yazisinda verinin kaynagini teyit etmedigini
bildirmisti; ancak tablo, bolumun kendi resmi web sitesinde
(jeofizikmuhendislik.iuc.edu.tr/tr/content/egitim/ders-programlari)
yayinlanan, Bolum Baskani imzali 4 resmi ders programi belgesiyle
(2025-2026 Guz/Bahar Lisans + Guz/Bahar Yuksek Lisans-Doktora) tek tek
karsilastirilarak dogrulandi: ~30 ogretim uyesinden 29'unun toplam ders
sayisi birebir eslesti (Tezel: 9=9, Hisarli: 9=9 dahil). Tek net
uyusmazlik: Zeynep Cansu (Excel'de 1, resmi programda 2 ders). Zorunlu/
Secmeli (Z/S) etiketi Lisans belgelerinde var ama YL/Doktora belgelerinde
yok; bu yuzden ders SAYILARI dogrulanmis, ancak Z/S siniflandirmasi
YL/Doktora duzeyinde tam bagimsiz teyit edilememistir.

Kaynak: Jeofizik_Ders_Yuku_Dagilimi.xlsx (Karabulut tarafindan derlenmis;
icerigi resmi bolum web sitesindeki ders programlariyla buyuk olcude
dogrulanmistir, bkz. yukarida)
  - "Ana Tablo" sayfasi: ders yuku verileri (Toplam Z Ders sutunu)
  - "Anabim Dallarına göre" sayfasi: resmi Jeofizik Anabilim Dali uye listesi

SADECE resmi Jeofizik Anabilim Dali (Sismoloji, Uygulamali Jeofizik,
Yer Fizigi) uyesi olan Prof. Dr.'lar listelenir. Nurullah Hanilci ve
Aysel Ersoy, Jeoloji Muhendisligi Bolumu'nun kadrolu ogretim uyeleridir;
Jeofizik ogrencilerine disaridan (hizmet) dersi verdikleri icin Jeofizik
ogretim uyesi sayilmazlar ve bu listeye DAHIL EDILMEZLER.
"""
import openpyxl

XLSX_PATH = "../data/Jeofizik_Ders_Yuku_Dagilimi.xlsx"
ANA_TABLO = "Ana Tablo"
ANABILIM_DALI_SHEET = "Anabim Dallarına göre"

TR_MAP = str.maketrans({
    "İ": "i", "I": "i", "ı": "i",
    "Ş": "s", "ş": "s", "Ğ": "g", "ğ": "g",
    "Ü": "u", "ü": "u", "Ö": "o", "ö": "o", "Ç": "c", "ç": "c",
})


def normalize(name):
    """Soyadina (son kelime) gore anahtar uretir; kisaltilmis ilk isimler
    (orn. 'Z. Mümtaz HİSARLI' vs 'Zihni Mümtaz Hisarlı') tam ad eslesmesini
    bozdugundan sadece soyad kullanilir."""
    name = name.replace("Dr. Öğr. Üyesi (Doç. Dr.)", "")
    words = name.translate(TR_MAP).split()
    return words[-1].lower() if words else ""


wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

dali_ws = wb[ANABILIM_DALI_SHEET]
resmi_uyeler = {}  # normalize(isim) -> anabilim dali
for row in dali_ws.iter_rows(min_row=9, max_row=30, values_only=True):
    anabilim_dali, isim, _unvan = row[0], row[1], row[2]
    if anabilim_dali and isim:
        resmi_uyeler[normalize(isim)] = anabilim_dali

ws = wb[ANA_TABLO]
headers = [c.value for c in ws[1]]
idx_name = headers.index("Öğretim Üyesi")
idx_unvan = headers.index("Ünvan")
idx_toplam_z_ders = headers.index("Toplam Z Ders")
idx_toplam_z_kredi = headers.index("Toplam Z Kredi")

profs = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[idx_name]
    if not name or row[idx_unvan] != "Prof. Dr.":
        continue
    key = normalize(name)
    if key not in resmi_uyeler:
        continue  # Jeofizik anabilim dali uyesi degil (disaridan ders veren), dahil edilmez
    profs.append((name, resmi_uyeler[key], row[idx_toplam_z_ders], row[idx_toplam_z_kredi]))

profs.sort(key=lambda p: -p[2])

zorunlu_olan = [p for p in profs if p[2] > 0]
zorunlu_olmayan = [p for p in profs if p[2] == 0]

print(f"Resmi Jeofizik Anabilim Dali'ndaki toplam Prof. Dr. sayisi: {len(profs)}")
print(f"Zorunlu dersi olan (Toplam Z Ders > 0) Prof. Dr. sayisi: {len(zorunlu_olan)}\n")

for name, dali, z_ders, z_kredi in zorunlu_olan:
    print(f"- {name} [{dali}] -> {z_ders} zorunlu ders, {z_kredi} kredi")

print(f"\nZorunlu dersi olmayan Prof. Dr. sayisi: {len(zorunlu_olmayan)}")
for name, dali, z_ders, z_kredi in zorunlu_olmayan:
    print(f"- {name} [{dali}]")
