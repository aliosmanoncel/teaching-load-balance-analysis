"""
Jeofizik Ders Yuku Dagilimi - TUM ogretim uyeleri icin zorunlu ders durumu.

!!! VERI DOGRULAMA NOTU (14.07.2026 itibariyle guncellendi) !!!
Bu Excel tablosu, 13.07.2026 tarihinde Doc. Dr. Savas Karabulut tarafindan bolum e-posta listesi ve X (Twitter) uzerinden kamuya acik olarak paylasilmistir. Bolum
Baskanligi 07.07.2026 tarihli yazisinda verinin kaynagini teyit etmedigini
bildirmisti; ancak tablo, bolumun kendi resmi web sitesinde
(jeofizikmuhendislik.iuc.edu.tr/tr/content/egitim/ders-programlari)
yayinlanan, Bolum Baskani imzali 4 resmi ders programi belgesiyle
(2025-2026 Guz/Bahar Lisans + Guz/Bahar Yuksek Lisans-Doktora) tek tek
karsilastirilarak dogrulandi: ~30 ogretim uyesinden 29'unun toplam ders
sayisi birebir eslesti (Tezel: 9=9, Hisarli: 9=9 dahil). Tek net
uyusmazlik: Zeynep Cansu (Excel'de 1, resmi programda 2 ders). Zorunlu/
Secmeli (Z/S) etiketi Lisans belgelerinde var ama YL/Doktora belgelerinde
yok; bu yuzden ders SAYILARI dogrulanmis, ancak Z/S siniflandirmasi
YL/Doktora duzeyinde tam bagimsiz teyit edilememistir.

Kaynak: Jeofizik_Ders_Yuku_Dagilimi.xlsx (Karabulut tarafindan derlenmis;
icerigi resmi bolum web sitesindeki ders programlariyla buyuk olcude
dogrulanmistir, bkz. yukarida)
  - "Ana Tablo" sayfasi: ders yuku verileri (Toplam Z Ders sutunu)
  - "Anabim Dallarına göre" sayfasi: resmi Jeofizik Anabilim Dali uye listesi

Her ogretim uyesi icin Toplam Z Ders > 0 ise (+), == 0 ise (-) isaretlenir.
Resmi Jeofizik Anabilim Dali (Sismoloji, Uygulamali Jeofizik, Yer Fizigi)
listesinde olmayanlar "Diger Bolum (disaridan ders)" olarak etiketlenir.
"""
import openpyxl

XLSX_PATH = "../data/Jeofizik_Ders_Yuku_Dagilimi.xlsx"
ANA_TABLO = "Ana Tablo"
ANABILIM_DALI_SHEET = "Anabim Dallarına göre"

TR_MAP = str.maketrans({
    "İ": "i", "I": "i", "ı": "i",
    "Ş": "s", "ş": "s", "Ğ": "g", "ğ": "g",
    "Ü": "u", "ü": "u", "Ö": "o", "ö": "o", "Ç": "c", "ç": "c",
})


def normalize(name):
    """Soyadina (son kelime) gore anahtar uretir; kisaltilmis ilk isimler
    tam ad eslesmesini bozdugundan sadece soyad kullanilir."""
    name = name.replace("Dr. Öğr. Üyesi (Doç. Dr.)", "")
    words = name.translate(TR_MAP).split()
    return words[-1].lower() if words else ""


# Jeofizik anabilim dali listesinde olmayan ama bilinen ogretim uyelerinin
# gercek bolumu (kullanicidan gelen bilgiyle dogrulanmistir)
BILINEN_DIGER_BOLUM = {
    "hanilci": "Jeoloji Mühendisliği (dışarıdan ders)",
    "ersoy": "Jeoloji Mühendisliği (dışarıdan ders)",
}

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

dali_ws = wb[ANABILIM_DALI_SHEET]
resmi_uyeler = {}  # normalize(isim) -> anabilim dali
for row in dali_ws.iter_rows(min_row=9, max_row=30, values_only=True):
    anabilim_dali, isim, _unvan = row[0], row[1], row[2]
    if anabilim_dali and isim:
        resmi_uyeler[normalize(isim)] = anabilim_dali

ws = wb[ANA_TABLO]
headers = [c.value for c in ws[1]]
idx_name = headers.index("Öğretim Üyesi")
idx_unvan = headers.index("Ünvan")
idx_toplam_z_ders = headers.index("Toplam Z Ders")

herkes = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[idx_name]
    if not name:
        continue
    key = normalize(name)
    dali = resmi_uyeler.get(key) or BILINEN_DIGER_BOLUM.get(key, "Diğer Bölüm (dışarıdan ders)")
    herkes.append((name, row[idx_unvan], dali, row[idx_toplam_z_ders]))

# Unvan sirasi: Prof. Dr. -> Doç. Dr. -> Dr. Öğr. Üyesi -> Araş. Gör. Dr.
unvan_sira = {"Prof. Dr.": 0, "Doç. Dr.": 1, "Dr. Öğr. Üyesi": 2, "Araş. Gör. Dr.": 3}
herkes.sort(key=lambda p: (unvan_sira.get(p[1], 9), -p[3]))

var = [p for p in herkes if p[3] > 0]
yok = [p for p in herkes if p[3] == 0]

print(f"Toplam ogretim uyesi: {len(herkes)}")
print(f"Zorunlu dersi VAR (+): {len(var)}")
print(f"Zorunlu dersi YOK (-): {len(yok)}\n")

print("=" * 70)
for name, unvan, dali, z_ders in herkes:
    isaret = "+" if z_ders > 0 else "-"
    ders_bilgi = f"{z_ders} zorunlu ders" if z_ders > 0 else "zorunlu ders yok"
    print(f"[{isaret}] {name} ({unvan}) - {dali} - {ders_bilgi}")
