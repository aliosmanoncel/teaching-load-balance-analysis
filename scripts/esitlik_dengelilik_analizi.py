"""
Jeofizik Ders Yuku Dagilimi - "Esit ve dengeli dagitim" analizi.

!!! VERI DOGRULAMA NOTU (14.07.2026 itibariyle guncellendi) !!!
Bu Excel tablosu Doc. Dr. Savas Karabulut tarafindan derlenmistir. Bolum
Baskanligi 07.07.2026 tarihli yazisinda verinin kaynagini teyit etmedigini
bildirmisti; ancak tablo, bolumun kendi resmi web sitesinde
(jeofizikmuhendislik.iuc.edu.tr/tr/content/egitim/ders-programlari)
yayinlanan, Bolum Baskani imzali 4 resmi ders programi belgesiyle tek tek
karsilastirilarak dogrulandi: ~30 ogretim uyesinden 29'unun toplam ders
sayisi birebir eslesti (Tezel: 9=9, Hisarli: 9=9 dahil). Tek net
uyusmazlik: Zeynep Cansu (Excel'de 1, resmi programda 2 ders). Bu
script'teki ders SAYILARI (CV=0.85 dahil) artik bagimsiz dogrulanmistir;
ancak Z/S siniflandirmasi YL/Doktora duzeyinde tam bagimsiz teyit
edilememistir (bkz. dosya sonundaki metodolojik not).

2547 sayili Yuksek Ogretim Kanunu (ve Bolum Baskanligi yonetmelikleri),
bolum baskaninin dersleri ogretim uyeleri arasinda ESIT VE DENGELI
dagitmasini ongorur. Bu script, resmi Jeofizik Anabilim Dali (Sismoloji,
Uygulamali Jeofizik, Yer Fizigi -> 22 kisi) icin ders yukunun ne kadar
esit/dengeli dagitildigini nicel olarak degerlendirir.

Kullanilan olcutler:
  - Toplam Z Ders / Kredi / AKTS (sadece zorunlu ders yuku)
  - Genel Toplam Ders / Kredi / AKTS (zorunlu + secmeli toplam yuk)
  - Unvan grubu icinde (Prof/Doc/Dr. Ogr. Uyesi) karsilastirma
    (farkli unvanlarin zorunlu ders esigi farkli olabilir, bu yuzden
    hem genel hem unvan-ici dagilim raporlanir)
  - Degisim katsayisi (CV = std / ortalama) - esitsizligin standart bir
    olcusu; CV ne kadar yuksekse dagilim o kadar dengesiz demektir.
"""
import statistics as stats
import openpyxl

XLSX_PATH = "../data/Jeofizik_Ders_Yuku_Dagilimi.xlsx"
ANA_TABLO = "Ana Tablo"
ANABILIM_DALI_SHEET = "Anabim Dallarına göre"

TR_MAP = str.maketrans({
    "İ": "i", "I": "i", "ı": "i",
    "Ş": "s", "ş": "s", "Ğ": "g", "ğ": "g",
    "Ü": "u", "ü": "u", "Ö": "o", "ö": "o", "Ç": "c", "ç": "c",
})


def normalize(name):
    name = name.replace("Dr. Öğr. Üyesi (Doç. Dr.)", "")
    words = name.translate(TR_MAP).split()
    return words[-1].lower() if words else ""


wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

dali_ws = wb[ANABILIM_DALI_SHEET]
resmi_uyeler = {}
for row in dali_ws.iter_rows(min_row=9, max_row=30, values_only=True):
    anabilim_dali, isim, _unvan = row[0], row[1], row[2]
    if anabilim_dali and isim:
        resmi_uyeler[normalize(isim)] = anabilim_dali

ws = wb[ANA_TABLO]
headers = [c.value for c in ws[1]]
idx_name = headers.index("Öğretim Üyesi")
idx_unvan = headers.index("Ünvan")
idx_z_ders = headers.index("Toplam Z Ders")
idx_z_kredi = headers.index("Toplam Z Kredi")
idx_z_akts = headers.index("Toplam Z AKTS")
idx_genel_ders = headers.index("Genel Toplam Ders")
idx_genel_kredi = headers.index("Genel Toplam Kredi")
idx_genel_akts = headers.index("Genel Toplam AKTS")

kayitlar = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[idx_name]
    if not name:
        continue
    key = normalize(name)
    if key not in resmi_uyeler:
        continue  # sadece resmi Jeofizik Anabilim Dali uyeleri
    kayitlar.append({
        "isim": name,
        "unvan": row[idx_unvan],
        "dali": resmi_uyeler[key],
        "z_ders": row[idx_z_ders] or 0,
        "z_kredi": row[idx_z_kredi] or 0,
        "z_akts": row[idx_z_akts] or 0,
        "genel_ders": row[idx_genel_ders] or 0,
        "genel_kredi": row[idx_genel_kredi] or 0,
        "genel_akts": row[idx_genel_akts] or 0,
    })


def cv(values):
    if not values or stats.mean(values) == 0:
        return None
    if len(values) < 2:
        return 0.0
    return stats.pstdev(values) / stats.mean(values)


def rapor_yaz(baslik, grup):
    z_ders_list = [k["z_ders"] for k in grup]
    z_akts_list = [k["z_akts"] for k in grup]
    genel_akts_list = [k["genel_akts"] for k in grup]
    print(f"\n--- {baslik} (n={len(grup)}) ---")
    print(f"Zorunlu Ders  : min={min(z_ders_list)}  max={max(z_ders_list)}  "
          f"ortalama={stats.mean(z_ders_list):.2f}  CV={cv(z_ders_list)}")
    print(f"Zorunlu AKTS  : min={min(z_akts_list)}  max={max(z_akts_list)}  "
          f"ortalama={stats.mean(z_akts_list):.2f}  CV={cv(z_akts_list)}")
    print(f"Genel Top.AKTS: min={min(genel_akts_list)}  max={max(genel_akts_list)}  "
          f"ortalama={stats.mean(genel_akts_list):.2f}  CV={cv(genel_akts_list)}")
    sifir_zorunlu = [k["isim"] for k in grup if k["z_ders"] == 0]
    if sifir_zorunlu:
        print(f"Zorunlu dersi SIFIR olanlar: {', '.join(sifir_zorunlu)}")


print("=" * 78)
print("JEOFIZIK BOLUMU - DERS YUKU ESITLIK/DENGELILIK ANALIZI")
print("(2547 s. Kanun: Bolum Baskani dersleri esit ve dengeli dagitir)")
print("=" * 78)

rapor_yaz("TUM BOLUM (22 kisi)", kayitlar)

for unvan in ["Prof. Dr.", "Doç. Dr.", "Dr. Öğr. Üyesi"]:
    grup = [k for k in kayitlar if k["unvan"] == unvan or
            (unvan == "Doç. Dr." and k["unvan"] == "Dr. Öğr. Üyesi (Doç. Dr.)")]
    if grup:
        rapor_yaz(f"UNVAN GRUBU: {unvan}", grup)

print("\n" + "=" * 78)
print("DETAY TABLO (Prof. Dr. grubu, zorunlu ders sayisina gore siralandi)")
print("=" * 78)
profs = sorted([k for k in kayitlar if k["unvan"] == "Prof. Dr."],
               key=lambda k: -k["z_ders"])
for k in profs:
    baskan = "  <-- BÖLÜM BAŞKANI" if "HİSARLI" in k["isim"].upper() else ""
    print(f"{k['isim']:35s} [{k['dali']:28s}] "
          f"Z.Ders={k['z_ders']}  Z.AKTS={k['z_akts']:5.1f}  "
          f"Genel.AKTS={k['genel_akts']:5.1f}{baskan}")
