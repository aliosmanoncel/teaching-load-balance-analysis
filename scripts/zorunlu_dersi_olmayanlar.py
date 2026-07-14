"""
Jeofizik Ders Yuku Dagilimi - Zorunlu dersi olmayan ogretim uyelerini bulur.

!!! VERI DOGRULAMA NOTU (14.07.2026 itibariyle guncellendi) !!!
Bu Excel tablosu Doc. Dr. Savas Karabulut tarafindan derlenmistir. Bolum
Baskanligi 07.07.2026 tarihli yazisinda verinin kaynagini teyit etmedigini
bildirmisti; ancak tablo, bolumun kendi resmi web sitesinde
(jeofizikmuhendislik.iuc.edu.tr/tr/content/egitim/ders-programlari)
yayinlanan, Bolum Baskani imzali 4 resmi ders programi belgesiyle
(2025-2026 Guz/Bahar Lisans + Guz/Bahar Yuksek Lisans-Doktora) tek tek
karsilastirilarak dogrulandi: ~30 ogretim uyesinden 29'unun toplam ders
sayisi birebir eslesti (Tezel: 9=9, Hisarli: 9=9 dahil). Tek net
uyusmazlik: Zeynep Cansu (Excel'de 1, resmi programda 2 ders). Zorunlu/
Secmeli (Z/S) etiketi Lisans belgelerinde var ama YL/Doktora belgelerinde
yok; bu yuzden ders SAYILARI dogrulanmis, ancak Z/S siniflandirmasi
YL/Doktora duzeyinde tam bagimsiz teyit edilememistir.

Kaynak: Jeofizik_Ders_Yuku_Dagilimi.xlsx (Karabulut tarafindan derlenmis;
icerigi resmi bolum web sitesindeki ders programlariyla buyuk olcude
dogrulanmistir, bkz. yukarida)
  - "Ana Tablo" sayfasi: ders yuku verileri (Toplam Z Ders == 0 kriteri)
  - "Anabim Dallarına göre" sayfasi: resmi Jeofizik Anabilim Dali uye listesi
    (Sismoloji, Uygulamali Jeofizik, Yer Fizigi -> 22 kisi)

Ana Tablo'da Jeofizik anabilim dallarina bagli olmayan, baska bolumden
hizmet dersi veren ogretim uyeleri de gorunebiliyor (orn. Jeoloji).
Bu yuzden sonuc, resmi anabilim dali listesiyle kesistirilir.
"""
import openpyxl

XLSX_PATH = "../data/Jeofizik_Ders_Yuku_Dagilimi.xlsx"
ANA_TABLO = "Ana Tablo"
ANABILIM_DALI_SHEET = "Anabim Dallarına göre"


TR_MAP = str.maketrans({
    "İ": "i", "I": "i", "ı": "i",
    "Ş": "s", "ş": "s", "Ğ": "g", "ğ": "g",
    "Ü": "u", "ü": "u", "Ö": "o", "ö": "o", "Ç": "c", "ç": "c",
})


def normalize(name):
    """Soyadina (son kelime) gore anahtar uretir; kisaltilmis ilk isimler
    (orn. 'Z. Mümtaz HİSARLI' vs 'Zihni Mümtaz Hisarlı') tam ad eslesmesini
    bozdugundan sadece soyad kullanilir."""
    name = name.replace("Dr. Öğr. Üyesi (Doç. Dr.)", "")
    words = name.translate(TR_MAP).split()
    return words[-1].lower() if words else ""


wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# Resmi Jeofizik Anabilim Dali uyeleri (unvan onekleri farkli yazilabiliyor,
# bu yuzden isim kismina gore normalize edip karsilastiriyoruz)
dali_ws = wb[ANABILIM_DALI_SHEET]
resmi_uyeler = set()
for row in dali_ws.iter_rows(min_row=9, max_row=30, values_only=True):
    anabilim_dali, isim, _unvan = row[0], row[1], row[2]
    if anabilim_dali and isim:
        resmi_uyeler.add(normalize(isim))

ws = wb[ANA_TABLO]
headers = [c.value for c in ws[1]]
idx_name = headers.index("Öğretim Üyesi")
idx_unvan = headers.index("Ünvan")
idx_toplam_z_ders = headers.index("Toplam Z Ders")

zorunlu_olmayanlar = []
haric_tutulanlar = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[idx_name]
    if not name:
        continue
    if row[idx_toplam_z_ders] == 0:
        if normalize(name) in resmi_uyeler:
            zorunlu_olmayanlar.append((name, row[idx_unvan]))
        else:
            haric_tutulanlar.append((name, row[idx_unvan]))

print(f"Zorunlu dersi olmayan Jeofizik ogretim uyesi sayisi: {len(zorunlu_olmayanlar)}\n")
for name, unvan in zorunlu_olmayanlar:
    print(f"- {name} ({unvan})")

print(f"\nHaric tutulanlar (resmi Jeofizik Anabilim Dali listesinde yok, "
      f"baska bolumden hizmet dersi veriyor olabilir): {len(haric_tutulanlar)}")
for name, unvan in haric_tutulanlar:
    print(f"- {name} ({unvan})")
